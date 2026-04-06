import streamlit as st
import pandas as pd
import requests
from datetime import datetime
import io

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Import Duty Calculator",
    page_icon="🛃",
    layout="wide",
)

# ── Load CSS ──────────────────────────────────────────────────────────────────
with open("style.css") as f:
    st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)

# ── Load commodity data ───────────────────────────────────────────────────────
@st.cache_data
def load_commodities():
    df = pd.read_csv("commodities.csv", dtype=str)
    df.columns = df.columns.str.strip()
    df["duty_pct"] = df["duty_pct"].str.replace(",", ".").astype(float)
    return df

# ── Fetch exchange rates from InforEuro API (official EC rates) ───────────────
@st.cache_data(ttl=3600)
def fetch_exchange_rates():
    """
    InforEuro official monthly EC accounting rates.
    Endpoint: https://ec.europa.eu/budg/inforeuro/api/public/currencies/
    Falls back to Frankfurter (ECB daily) if InforEuro is unavailable.
    """
    rates = {}
    try:
        url = "https://ec.europa.eu/budg/inforeuro/api/public/monthly-rates"
        resp = requests.get(url, timeout=8)
        if resp.status_code == 200:
            data = resp.json()
            for item in data:
                iso = item.get("isoA3Code", "").upper()
                rate = item.get("value")
                if iso and rate:
                    rates[iso] = float(rate)
            if rates:
                return rates, "InforEuro (EC officieel)"
    except Exception:
        pass

    # Fallback: Frankfurter (ECB daily reference rates, EUR base)
    try:
        resp = requests.get("https://api.frankfurter.app/latest?from=EUR", timeout=8)
        if resp.status_code == 200:
            data = resp.json()
            # rates are "how many X per 1 EUR", we need "how many EUR per 1 X"
            for currency, rate in data.get("rates", {}).items():
                rates[currency.upper()] = round(1 / rate, 6)
            rates["EUR"] = 1.0
            return rates, "Frankfurter / ECB (dagelijks)"
    except Exception:
        pass

    return {}, "Niet beschikbaar"

# ── Session state init ────────────────────────────────────────────────────────
def init_state():
    if "lines" not in st.session_state:
        st.session_state.lines = [new_line()]
    if "ref" not in st.session_state:
        st.session_state.ref = ""
    if "user" not in st.session_state:
        st.session_state.user = ""

def new_line():
    return {
        "commodity_idx": None,
        "currency": "USD",
        "invoice_value": 0.0,
        "manual_rate": None,   # None = use live rate
    }

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    init_state()
    commodities_df = load_commodities()
    commodity_options = [
        f"{row['commodity_code']} – {row['description']}"
        for _, row in commodities_df.iterrows()
    ]

    exchange_rates, rate_source = fetch_exchange_rates()

    # ── Header ────────────────────────────────────────────────────────────────
    st.markdown('<div class="app-header"><span class="flag">🛃</span> Import Duty Calculator</div>', unsafe_allow_html=True)

    # ── Reference & user ─────────────────────────────────────────────────────
    col_ref, col_user, col_rate_info = st.columns([2, 2, 3])
    with col_ref:
        st.session_state.ref = st.text_input("📋 Commerciële referentie", value=st.session_state.ref, placeholder="bv. INV-2025-001")
    with col_user:
        st.session_state.user = st.text_input("👤 Gebruiker", value=st.session_state.user, placeholder="Naam")
    with col_rate_info:
        if exchange_rates:
            st.markdown(
                f'<div class="rate-badge">✅ Wisselkoersen geladen via <b>{rate_source}</b> '
                f'&nbsp;|&nbsp; {datetime.now().strftime("%d/%m/%Y %H:%M")}</div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown('<div class="rate-badge warn">⚠️ Wisselkoersen niet beschikbaar – gebruik manuele invoer</div>', unsafe_allow_html=True)

    st.divider()

    # ── Lines table header ────────────────────────────────────────────────────
    col_widths = [3, 1.2, 1.4, 1.4, 1.2, 1.2, 1.2, 1.2, 0.5]
    headers = ["Commodity", "Munt", "Factuurwaarde", "Koers → EUR", "Waarde EUR", "Duty %", "Duty", "BTW (21%)", ""]
    h_cols = st.columns(col_widths)
    for hc, ht in zip(h_cols, headers):
        hc.markdown(f"<div class='col-header'>{ht}</div>", unsafe_allow_html=True)

    # ── Per-line inputs ───────────────────────────────────────────────────────
    lines_data = []   # computed results per line

    for i, line in enumerate(st.session_state.lines):
        c = st.columns(col_widths)

        # Commodity selectbox with search
        with c[0]:
            sel = st.selectbox(
                f"commodity_{i}", commodity_options,
                index=line["commodity_idx"] if line["commodity_idx"] is not None else 0,
                key=f"comm_{i}",
                label_visibility="collapsed",
            )
            line["commodity_idx"] = commodity_options.index(sel)
            comm_code = sel.split(" – ")[0]
            row = commodities_df[commodities_df["commodity_code"] == comm_code].iloc[0]
            duty_pct = row["duty_pct"]

        # Currency
        common_currencies = ["USD", "GBP", "CHF", "CNY", "JPY", "CAD", "AUD", "NOK", "SEK", "DKK", "EUR"]
        with c[1]:
            currency = st.selectbox(
                f"curr_{i}", common_currencies,
                index=common_currencies.index(line["currency"]) if line["currency"] in common_currencies else 0,
                key=f"cur_{i}",
                label_visibility="collapsed",
            )
            line["currency"] = currency

        # Invoice value
        with c[2]:
            inv_val = st.number_input(
                f"inv_{i}", min_value=0.0, value=float(line["invoice_value"]),
                step=100.0, format="%.2f",
                key=f"inv_{i}",
                label_visibility="collapsed",
            )
            line["invoice_value"] = inv_val

        # Exchange rate (live or manual)
        live_rate = exchange_rates.get(currency, None)
        if currency == "EUR":
            live_rate = 1.0
        with c[3]:
            default_rate = line["manual_rate"] if line["manual_rate"] is not None else (live_rate or 1.0)
            exch_rate = st.number_input(
                f"rate_{i}", min_value=0.0, value=float(default_rate),
                step=0.0001, format="%.4f",
                key=f"rate_{i}",
                label_visibility="collapsed",
                help="Automatisch via InforEuro/ECB. Pas manueel aan indien gewenst.",
            )
            line["manual_rate"] = exch_rate

        # Calculations
        value_eur = inv_val * exch_rate
        duty_calc = value_eur * (duty_pct / 100)
        vat_calc = (value_eur + duty_calc) * 0.21
        total_taxes = duty_calc + vat_calc

        with c[4]:
            st.markdown(f"<div class='calc-cell'>€ {value_eur:,.2f}</div>", unsafe_allow_html=True)
        with c[5]:
            st.markdown(f"<div class='calc-cell'>{duty_pct:.2f}%</div>", unsafe_allow_html=True)
        with c[6]:
            st.markdown(f"<div class='calc-cell'>€ {duty_calc:,.2f}</div>", unsafe_allow_html=True)
        with c[7]:
            st.markdown(f"<div class='calc-cell'>€ {vat_calc:,.2f}</div>", unsafe_allow_html=True)
        with c[8]:
            if st.button("🗑", key=f"del_{i}", help="Verwijder lijn", disabled=len(st.session_state.lines) == 1):
                st.session_state.lines.pop(i)
                st.rerun()

        lines_data.append({
            "Commodity": sel,
            "Munt": currency,
            "Factuurwaarde": inv_val,
            "Koers": exch_rate,
            "Waarde EUR": value_eur,
            "Duty %": duty_pct,
            "Duty": duty_calc,
            "BTW (21%)": vat_calc,
            "Totaal Taxes": total_taxes,
        })

    # ── Add line ──────────────────────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    col_add, col_spacer = st.columns([1, 8])
    with col_add:
        if st.button("➕ Lijn toevoegen", use_container_width=True):
            st.session_state.lines.append(new_line())
            st.rerun()

    st.divider()

    # ── Totals ────────────────────────────────────────────────────────────────
    if lines_data:
        df_lines = pd.DataFrame(lines_data)
        total_eur = df_lines["Waarde EUR"].sum()
        total_duty = df_lines["Duty"].sum()
        total_vat = df_lines["BTW (21%)"].sum()
        total_taxes = df_lines["Totaal Taxes"].sum()

        st.markdown("### 📊 Totalen")
        t1, t2, t3, t4 = st.columns(4)
        t1.metric("Totaal Factuurwaarde (EUR)", f"€ {total_eur:,.2f}")
        t2.metric("Totaal Douanerechten", f"€ {total_duty:,.2f}")
        t3.metric("Totaal BTW (21%)", f"€ {total_vat:,.2f}")
        t4.metric("Totaal Belastingen", f"€ {total_taxes:,.2f}")

        st.divider()

        # ── Export ────────────────────────────────────────────────────────────
        st.markdown("### 📥 Export")
        export_df = df_lines.copy()
        export_df.insert(0, "Referentie", st.session_state.ref)
        export_df.insert(1, "Gebruiker", st.session_state.user)

        # Excel export
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            export_df.to_excel(writer, index=False, sheet_name="Berekening")
            # Totals row
            ws = writer.sheets["Berekening"]
            from openpyxl.styles import Font, PatternFill
            last_row = len(export_df) + 2
            ws.cell(last_row, 1, "TOTAAL").font = Font(bold=True)
            ws.cell(last_row, 7, total_eur).font = Font(bold=True)
            ws.cell(last_row, 9, total_duty).font = Font(bold=True)
            ws.cell(last_row, 10, total_vat).font = Font(bold=True)
            ws.cell(last_row, 11, total_taxes).font = Font(bold=True)

        col_xl, col_csv = st.columns([1, 1])
        with col_xl:
            st.download_button(
                "⬇️ Download Excel",
                data=buffer.getvalue(),
                file_name=f"duty_calc_{st.session_state.ref or 'export'}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with col_csv:
            csv_bytes = export_df.to_csv(index=False, sep=";", decimal=",").encode("utf-8-sig")
            st.download_button(
                "⬇️ Download CSV",
                data=csv_bytes,
                file_name=f"duty_calc_{st.session_state.ref or 'export'}_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True,
            )

    # ── Footer ────────────────────────────────────────────────────────────────
    st.markdown(
        '<div class="footer">Berekeningen zijn indicatief. Controleer altijd de officiële TARIC-tarieven. '
        f'Wisselkoersbron: {rate_source}.</div>',
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    main()

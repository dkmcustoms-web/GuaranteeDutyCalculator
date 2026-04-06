import os
import streamlit as st
import pandas as pd
import requests
import base64
import io
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable, Image as RLImage,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Guarantee Calculation – DKM",
    page_icon="🛃",
    layout="wide",
)

# ── Load CSS ──────────────────────────────────────────────────────────────────
with open("style.css") as f:
    st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)

# ── Load commodity data ───────────────────────────────────────────────────────
def _csv_mtime():
    """Return file modification time so cache invalidates when CSV changes."""
    try:
        return os.path.getmtime("commodities.csv")
    except Exception:
        return 0

@st.cache_data
def load_commodities(_mtime=None):
    df = pd.read_csv("commodities.csv", dtype=str)
    df.columns = df.columns.str.strip()
    df["duty_pct"] = df["duty_pct"].str.replace(",", ".").astype(float)
    df["commodity_code"] = df["commodity_code"].str.strip()
    return df

# ── Fetch exchange rates ──────────────────────────────────────────────────────
@st.cache_data(ttl=3600)
def fetch_exchange_rates():
    rates = {}
    try:
        url = "https://ec.europa.eu/budg/inforeuro/api/public/monthly-rates"
        resp = requests.get(url, timeout=8)
        if resp.status_code == 200:
            for item in resp.json():
                iso = item.get("isoA3Code", "").upper()
                rate = item.get("value")
                if iso and rate:
                    rates[iso] = float(rate)
            if rates:
                return rates, "InforEuro (EC officieel)"
    except Exception:
        pass
    try:
        resp = requests.get("https://api.frankfurter.app/latest?from=EUR", timeout=8)
        if resp.status_code == 200:
            for currency, rate in resp.json().get("rates", {}).items():
                rates[currency.upper()] = round(1 / rate, 6)
            rates["EUR"] = 1.0
            return rates, "Frankfurter / ECB (dagelijks)"
    except Exception:
        pass
    return {}, "Niet beschikbaar"

# ── Logo helpers (same pattern as working DKM apps) ──────────────────────────
LOGO_FILENAME = "DKM-Logo-kleur-1024x276.png"
LOGO_PATHS = [
    LOGO_FILENAME,
    os.path.join(os.path.dirname(os.path.abspath(__file__)), LOGO_FILENAME),
    f"/mnt/user-data/uploads/{LOGO_FILENAME}",
    f"static/{LOGO_FILENAME}",
]

def _find_logo() -> str:
    """Return first existing logo path, or empty string."""
    for path in LOGO_PATHS:
        if os.path.exists(path):
            return path
    return ""

def _logo_b64_for_pdf() -> tuple:
    """Return (base64_str, path) for PDF embedding."""
    path = _find_logo()
    if not path:
        return None, None
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode(), path
    except Exception:
        return None, None

# ── Session state init ────────────────────────────────────────────────────────
def init_state():
    defaults = {
        "lines": [new_line()],
        "ref": "",
        "user": "",
        "currency": None,   # single currency for whole dossier
        "manual_rate": None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

def new_line():
    return {"typed_code": "", "commodity_label": "", "invoice_value": 0.0}

# ── PDF generation ────────────────────────────────────────────────────────────
def build_pdf(lines_data, ref, user, currency, rate, rate_source,
              total_eur, total_duty, total_vat, total_taxes, logo_path=None):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=15*mm, bottomMargin=15*mm,
    )
    DKM_RED = colors.HexColor("#D94F2B")
    LIGHT_BLUE = colors.HexColor("#f0f4fa")
    DARK = colors.HexColor("#1a2e4a")
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("title", fontSize=16, textColor=DKM_RED,
                                 fontName="Helvetica-Bold", spaceAfter=2)
    sub_style = ParagraphStyle("sub", fontSize=9, textColor=DARK,
                               fontName="Helvetica", spaceAfter=1)
    small_style = ParagraphStyle("small", fontSize=7.5, textColor=colors.grey,
                                 fontName="Helvetica")
    cell_style = ParagraphStyle("cell", fontSize=8, fontName="Helvetica",
                                leading=10, wordWrap='CJK')

    story = []

    # ── Header: logo left, title right ───────────────────────────────────────
    _script_dir = os.path.dirname(os.path.abspath(__file__))
    _logo_candidates = [
        logo_path,
        os.path.join(_script_dir, "DKM-Logo-kleur-1024x276.png"),
        "DKM-Logo-kleur-1024x276.png",
    ]
    logo = None
    for _candidate in _logo_candidates:
        if _candidate and os.path.exists(str(_candidate)):
            try:
                logo = RLImage(_candidate, width=45*mm, height=12*mm)
                break
            except Exception:
                continue
    if logo is None:
        logo = Paragraph("<b>DKM</b>", ParagraphStyle(
            "lf", fontSize=14, textColor=DKM_RED, fontName="Helvetica-Bold"))

    header_data = [[
        logo,
        Paragraph("Guarantee Calculation", title_style),
    ]]
    header_tbl = Table(header_data, colWidths=[55*mm, None])
    header_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("LINEBELOW", (0, 0), (-1, 0), 1.5, DKM_RED),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
    ]))
    story.append(header_tbl)
    story.append(Spacer(1, 5*mm))

    # ── Meta info ─────────────────────────────────────────────────────────────
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    meta_data = [
        [Paragraph(f"<b>Referentie:</b> {ref or '–'}", sub_style),
         Paragraph(f"<b>Gebruiker:</b> {user or '–'}", sub_style),
         Paragraph(f"<b>Munt:</b> {currency or '–'}  |  <b>Koers:</b> {rate:.4f}", sub_style),
         Paragraph(f"<b>Datum:</b> {now}", sub_style)],
    ]
    meta_tbl = Table(meta_data, colWidths=[None, None, None, None])
    meta_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT_BLUE),
        ("ROUNDEDCORNERS", [4]),
        ("PADDING", (0, 0), (-1, -1), 5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(meta_tbl)
    story.append(Spacer(1, 5*mm))

    # ── Lines table ───────────────────────────────────────────────────────────
    col_headers = [
        "GN-code", "Omschrijving", "Factuurwaarde",
        "Waarde EUR", "Duty %", "Duty", "BTW (21%)", "Totaal Taxes"
    ]
    col_w = [22*mm, 55*mm, 22*mm, 22*mm, 14*mm, 20*mm, 20*mm, 20*mm]

    header_row = [Paragraph(f"<b>{h}</b>", ParagraphStyle(
        "th", fontSize=8, fontName="Helvetica-Bold",
        textColor=colors.white, alignment=TA_CENTER)) for h in col_headers]

    rows = [header_row]
    for ld in lines_data:
        comm_parts = ld["Commodity"].split(" – ", 1)
        code = comm_parts[0].strip()
        desc = comm_parts[1].strip() if len(comm_parts) > 1 else ""
        rows.append([
            Paragraph(code, cell_style),
            Paragraph(desc, cell_style),
            Paragraph(f"{ld['Factuurwaarde']:,.2f}", ParagraphStyle("r", fontSize=8, fontName="Helvetica", alignment=TA_RIGHT)),
            Paragraph(f"€ {ld['Waarde EUR']:,.2f}", ParagraphStyle("r", fontSize=8, fontName="Helvetica", alignment=TA_RIGHT)),
            Paragraph(f"{ld['Duty %']:.2f}%", ParagraphStyle("r", fontSize=8, fontName="Helvetica", alignment=TA_RIGHT)),
            Paragraph(f"€ {ld['Duty']:,.2f}", ParagraphStyle("r", fontSize=8, fontName="Helvetica", alignment=TA_RIGHT)),
            Paragraph(f"€ {ld['BTW (21%)']:,.2f}", ParagraphStyle("r", fontSize=8, fontName="Helvetica", alignment=TA_RIGHT)),
            Paragraph(f"€ {ld['Totaal Taxes']:,.2f}", ParagraphStyle("r", fontSize=8, fontName="Helvetica", alignment=TA_RIGHT)),
        ])

    # Totals row
    bold_r = ParagraphStyle("br", fontSize=8, fontName="Helvetica-Bold", alignment=TA_RIGHT)
    rows.append([
        Paragraph("<b>TOTAAL</b>", ParagraphStyle("bl", fontSize=8, fontName="Helvetica-Bold")),
        Paragraph("", cell_style),
        Paragraph("", cell_style),
        Paragraph(f"<b>€ {total_eur:,.2f}</b>", bold_r),
        Paragraph("", cell_style),
        Paragraph(f"<b>€ {total_duty:,.2f}</b>", bold_r),
        Paragraph(f"<b>€ {total_vat:,.2f}</b>", bold_r),
        Paragraph(f"<b>€ {total_taxes:,.2f}</b>", bold_r),
    ])

    tbl = Table(rows, colWidths=col_w, repeatRows=1)
    n = len(rows)
    tbl.setStyle(TableStyle([
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), DKM_RED),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 5),
        ("TOPPADDING", (0, 0), (-1, 0), 5),
        # Data rows
        ("FONTSIZE", (0, 1), (-1, n-2), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, n-2), [colors.white, LIGHT_BLUE]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 1), (-1, n-2), 4),
        ("BOTTOMPADDING", (0, 1), (-1, n-2), 4),
        # Totals row
        ("BACKGROUND", (0, n-1), (-1, n-1), colors.HexColor("#dce6f0")),
        ("TOPPADDING", (0, n-1), (-1, n-1), 5),
        ("BOTTOMPADDING", (0, n-1), (-1, n-1), 5),
        # Grid
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#c8d6e8")),
        ("LINEBELOW", (0, 0), (-1, 0), 1.5, DKM_RED),
        ("LINEABOVE", (0, n-1), (-1, n-1), 1, DARK),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 6*mm))

    # ── Summary boxes ─────────────────────────────────────────────────────────
    sum_data = [[
        Paragraph(f"<b>Totaal Waarde EUR</b><br/>€ {total_eur:,.2f}", sub_style),
        Paragraph(f"<b>Totaal Douanerechten</b><br/>€ {total_duty:,.2f}", sub_style),
        Paragraph(f"<b>Totaal BTW (21%)</b><br/>€ {total_vat:,.2f}", sub_style),
        Paragraph(f"<b>Totaal Belastingen</b><br/>€ {total_taxes:,.2f}",
                  ParagraphStyle("sum_hi", fontSize=9, fontName="Helvetica-Bold",
                                 textColor=DKM_RED)),
    ]]
    sum_tbl = Table(sum_data, colWidths=[None]*4)
    sum_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (2, 0), LIGHT_BLUE),
        ("BACKGROUND", (3, 0), (3, 0), colors.HexColor("#fde8e2")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#c8d6e8")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#c8d6e8")),
        ("PADDING", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(sum_tbl)
    story.append(Spacer(1, 8*mm))

    # ── Footer ────────────────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey))
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        f"Wisselkoersbron: {rate_source} | "
        "Berekeningen zijn indicatief. Controleer altijd de officiële TARIC-tarieven. "
        f"Gegenereerd op {now} door DKM Guarantee Calculation App.",
        small_style,
    ))

    doc.build(story)
    buf.seek(0)
    return buf.read()

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    init_state()
    commodities_df = load_commodities(_mtime=_csv_mtime())
    exchange_rates, rate_source = fetch_exchange_rates()

    COMMON_CURRENCIES = ["USD", "GBP", "CHF", "CNY", "JPY", "CAD", "AUD",
                         "NOK", "SEK", "DKK", "EUR", "HKD", "SGD", "INR"]

    # ── Header: red banner with logo (st.image) + title ─────────────────────
    logo_path = _find_logo()

    # Build banner: columns inside a styled container
    st.markdown('<div class="app-banner">', unsafe_allow_html=True)
    col_logo, col_title = st.columns([1, 4])
    with col_logo:
        if logo_path:
            st.image(logo_path, width=190)
        else:
            st.markdown('<span style="color:white;font-size:1.8rem;font-weight:900;">DKM</span>',
                        unsafe_allow_html=True)
    with col_title:
        st.markdown('<div class="banner-title">Guarantee Calculation</div>',
                    unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    # Reset button
    _, col_reset = st.columns([9, 1])
    with col_reset:
        if st.button("🔄 Reset", help="Wis alle gegevens en start opnieuw", use_container_width=True):
            for key in ["lines", "ref", "user", "currency", "manual_rate",
                        "currency_select", "global_rate"]:
                if key in st.session_state:
                    del st.session_state[key]
            st.rerun()

    # ── Dossier info row ──────────────────────────────────────────────────────
    col_ref, col_user, col_cur, col_rate, col_badge = st.columns([2, 2, 1.4, 1.4, 2.5])
    with col_ref:
        st.session_state.ref = st.text_input(
            "📋 Commerciële referentie", value=st.session_state.ref,
            placeholder="bv. INV-2025-001")
    with col_user:
        st.session_state.user = st.text_input(
            "👤 Gebruiker", value=st.session_state.user, placeholder="Naam")
    with col_cur:
        cur_options = [None] + COMMON_CURRENCIES
        saved_cur = st.session_state.currency
        cur_idx = cur_options.index(saved_cur) if saved_cur in cur_options else 0
        chosen_currency = st.selectbox(
            "💱 Munteenheid dossier *",
            options=cur_options,
            index=cur_idx,
            format_func=lambda x: "— kies munteenheid —" if x is None else x,
            key="currency_select")
        st.session_state.currency = chosen_currency
    with col_rate:
        live_rate = 1.0 if chosen_currency == "EUR" else (
            exchange_rates.get(chosen_currency, 1.0) if chosen_currency else 1.0)
        default_rate = st.session_state.manual_rate if st.session_state.manual_rate is not None else live_rate
        exch_rate = st.number_input(
            "Koers → EUR", min_value=0.00001, value=float(default_rate),
            step=0.0001, format="%.4f", key="global_rate",
            help="Automatisch via InforEuro/ECB. Pas manueel aan indien gewenst.",
            disabled=(chosen_currency is None))
        st.session_state.manual_rate = exch_rate
    with col_badge:
        if exchange_rates:
            st.markdown(
                f'<div class="rate-badge">✅ <b>{rate_source}</b><br>'
                f'{datetime.now().strftime("%d/%m/%Y %H:%M")}</div>',
                unsafe_allow_html=True)
        else:
            st.markdown(
                '<div class="rate-badge warn">⚠️ Wisselkoersen niet beschikbaar</div>',
                unsafe_allow_html=True)

    st.divider()

    # ── Block lines if no currency chosen ─────────────────────────────────────
    if chosen_currency is None:
        st.warning("⚠️ Kies eerst een munteenheid voordat je goederencodes kan invullen.")
        st.stop()

    # ── Column headers ────────────────────────────────────────────────────────
    col_widths = [1.5, 1.8, 1.5, 1.3, 1.3, 1.3, 1.3, 0.4]
    headers = ["GN-code (8 cijfers)", "Omschrijving",
               f"Factuurwaarde ({chosen_currency})",
               "Waarde EUR", "Duty %", "Duty", "BTW (21%)", ""]
    h_cols = st.columns(col_widths)
    for hc, ht in zip(h_cols, headers):
        hc.markdown(f"<div class='col-header'>{ht}</div>", unsafe_allow_html=True)

    # ── Lines ─────────────────────────────────────────────────────────────────
    lines_data = []

    for i, line in enumerate(st.session_state.lines):
        c = st.columns(col_widths)

        # GN-code: user types 8-digit code → lookup → show description or error
        # Determine status first so we can colour the whole row
        code_raw = st.session_state.get(f"gncode_{i}", line.get("typed_code", ""))
        code = code_raw.strip()
        if code and len(code) == 8:
            row_m = commodities_df[commodities_df["commodity_code"] == code]
            code_status = "found" if not row_m.empty else "notfound"
        elif code:
            row_m = pd.DataFrame()
            code_status = "partial"
        else:
            row_m = pd.DataFrame()
            code_status = "empty"

        # Row background: red if code not found
        row_bg = " row-error" if code_status == "notfound" else ""

        with c[0]:
            typed_code = st.text_input(
                f"gncode_{i}",
                value=line.get("typed_code", ""),
                placeholder="bv. 39269097",
                key=f"gncode_{i}",
                label_visibility="collapsed",
                max_chars=8,
            )
            line["typed_code"] = typed_code.strip()
            if code_status == "partial":
                st.caption(f"⏳ {len(code)}/8 cijfers")

            if code_status == "found":
                row_f = row_m.iloc[0]
                duty_pct = row_f["duty_pct"]
                commodity_label = f"{code} – {row_f['description']}"
                desc_display = row_f["description"]
            else:
                duty_pct = 0.0
                commodity_label = ""
                desc_display = ""

        # Omschrijving kolom
        with c[1]:
            if code_status == "found":
                st.markdown(f"<div class='calc-cell desc-cell{row_bg}'>{desc_display}</div>",
                            unsafe_allow_html=True)
            elif code_status == "notfound":
                st.markdown("<div class='calc-cell desc-cell row-error'>❌ Code niet gevonden</div>",
                            unsafe_allow_html=True)
            else:
                st.markdown("<div class='calc-cell desc-cell' style='color:#999;'>–</div>",
                            unsafe_allow_html=True)

        # Invoice value — auto-add new line on Enter (value change) if last line
        with c[2]:
            prev_val = float(line.get("invoice_value", 0.0))
            inv_val = st.number_input(
                f"inv_{i}", min_value=0.0, value=prev_val,
                step=100.0, format="%.2f",
                key=f"inv_{i}",
                label_visibility="collapsed",
            )
            line["invoice_value"] = inv_val
            # If value changed to > 0 and this is the last line → add new line
            is_last_line = (i == len(st.session_state.lines) - 1)
            if inv_val > 0 and inv_val != prev_val and is_last_line:
                st.session_state.lines.append(new_line())
                st.rerun()

        # Calculations
        value_eur = inv_val * exch_rate
        duty_calc = value_eur * (duty_pct / 100)
        vat_calc = (value_eur + duty_calc) * 0.21
        total_taxes = duty_calc + vat_calc

        with c[3]:
            st.markdown(f"<div class='calc-cell{row_bg}'>€ {value_eur:,.2f}</div>", unsafe_allow_html=True)
        with c[4]:
            st.markdown(f"<div class='calc-cell{row_bg}'>{duty_pct:.2f}%</div>", unsafe_allow_html=True)
        with c[5]:
            st.markdown(f"<div class='calc-cell{row_bg}'>€ {duty_calc:,.2f}</div>", unsafe_allow_html=True)
        with c[6]:
            st.markdown(f"<div class='calc-cell{row_bg}'>€ {vat_calc:,.2f}</div>", unsafe_allow_html=True)
        with c[7]:
            if st.button("🗑", key=f"del_{i}", help="Verwijder lijn",
                         disabled=len(st.session_state.lines) == 1):
                st.session_state.lines.pop(i)
                st.rerun()

        lines_data.append({
            "Commodity": commodity_label,
            "Munt": chosen_currency,
            "Factuurwaarde": inv_val,
            "Koers": exch_rate,
            "Waarde EUR": value_eur,
            "Duty %": duty_pct,
            "Duty": duty_calc,
            "BTW (21%)": vat_calc,
            "Totaal Taxes": total_taxes,
        })

    # ── Add line ──────────────────────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    col_add, _ = st.columns([1, 8])
    with col_add:
        if st.button("➕ Lijn toevoegen", use_container_width=True):
            st.session_state.lines.append(new_line())
            st.rerun()

    st.divider()

    # ── Totals ────────────────────────────────────────────────────────────────
    if lines_data:
        df_lines = pd.DataFrame(lines_data)
        total_eur = df_lines["Waarde EUR"].sum()
        total_duty = df_lines["Duty"].sum()
        total_vat = df_lines["BTW (21%)"].sum()
        total_taxes = df_lines["Totaal Taxes"].sum()

        st.markdown("### 📊 Totalen")
        t1, t2, t3, t4 = st.columns(4)
        t1.metric("Totaal Waarde EUR", f"€ {total_eur:,.2f}")
        t2.metric("Totaal Douanerechten", f"€ {total_duty:,.2f}")
        t3.metric("Totaal BTW (21%)", f"€ {total_vat:,.2f}")
        t4.metric("Totaal Belastingen", f"€ {total_taxes:,.2f}")

        st.divider()

        # ── Export ────────────────────────────────────────────────────────────
        st.markdown("### 📥 Export")
        export_df = df_lines.copy()
        export_df.insert(0, "Referentie", st.session_state.ref)
        export_df.insert(1, "Gebruiker", st.session_state.user)

        col_pdf, col_xl, col_csv = st.columns([1, 1, 1])

        # PDF — filter lege lijnen (geen commodity geselecteerd)
        pdf_lines = [ld for ld in lines_data if ld.get("Commodity")]
        with col_pdf:
            _, _pdf_logo_path = _logo_b64_for_pdf()
            pdf_bytes = build_pdf(
                pdf_lines,
                st.session_state.ref,
                st.session_state.user,
                chosen_currency,
                exch_rate,
                rate_source,
                total_eur, total_duty, total_vat, total_taxes,
                logo_path=_pdf_logo_path,
            )
            st.download_button(
                "🖨️ Download PDF",
                data=pdf_bytes,
                file_name=f"guarantee_{st.session_state.ref or 'export'}_{datetime.now().strftime('%Y%m%d')}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )

        # Excel
        with col_xl:
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                export_df.to_excel(writer, index=False, sheet_name="Berekening")
                from openpyxl.styles import Font
                ws = writer.sheets["Berekening"]
                last_row = len(export_df) + 2
                ws.cell(last_row, 1, "TOTAAL").font = Font(bold=True)
                ws.cell(last_row, 7, total_eur).font = Font(bold=True)
                ws.cell(last_row, 9, total_duty).font = Font(bold=True)
                ws.cell(last_row, 10, total_vat).font = Font(bold=True)
                ws.cell(last_row, 11, total_taxes).font = Font(bold=True)
            st.download_button(
                "⬇️ Download Excel",
                data=buffer.getvalue(),
                file_name=f"guarantee_{st.session_state.ref or 'export'}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        # CSV
        with col_csv:
            csv_bytes = export_df.to_csv(index=False, sep=";", decimal=",").encode("utf-8-sig")
            st.download_button(
                "⬇️ Download CSV",
                data=csv_bytes,
                file_name=f"guarantee_{st.session_state.ref or 'export'}_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True,
            )

    # ── Footer ────────────────────────────────────────────────────────────────
    st.markdown(
        f'<div class="footer">'
        f'Berekeningen zijn indicatief — controleer altijd de officiële TARIC-tarieven. '
        f'Wisselkoersbron: {rate_source}. &nbsp;|&nbsp; '
        f'<b>DKM-Customs</b> — Developed by Luc De Kerf</div>',
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    main()
